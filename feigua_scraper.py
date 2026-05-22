"""
飞瓜平台B站UP主数据抓取脚本
抓取路径: MCN机构列表 → 每个MCN的旗下达人 → 可选: UP主详情补充
支持断点续抓，中断后重新运行自动续抓

用法:
  1. 在 config.yaml 的 feigua_cookies 列表中填入 Cookie
  2. 运行: .venv/Scripts/python.exe feigua_scraper.py
"""

import requests
import time
import json
import os
import re
import yaml
from datetime import datetime
from tqdm import tqdm

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([os.path.join(".venv", "Scripts", "pip.exe"), "install", "openpyxl"])
    import openpyxl


# ============ 配置 ============
CONFIG_FILE = "config.yaml"
CHECKPOINT_DIR = "checkpoints"
MCN_CHECKPOINT = os.path.join(CHECKPOINT_DIR, "feigua_mcn.json")
BLOGGER_CHECKPOINT = os.path.join(CHECKPOINT_DIR, "feigua_blogger.json")
RANK_CHECKPOINT = os.path.join(CHECKPOINT_DIR, "feigua_rank.json")

BASE_URL = "https://bz.feigua.cn"
REQUEST_DELAY = 1.5  # 请求间隔（秒）
SAVE_EVERY = 5       # 每N个MCN保存一次断点


# ============ 工具函数 ============

def ensure_dir():
    os.makedirs(CHECKPOINT_DIR, exist_ok=True)


def load_config() -> dict:
    if not os.path.exists(CONFIG_FILE):
        print(f"错误: 找不到 {CONFIG_FILE}")
        exit(1)
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    cookies = cfg.get("feigua_cookies", [])
    if not cookies:
        print(f"错误: {CONFIG_FILE} 中未配置 feigua_cookies")
        exit(1)

    cookie = cookies[0].strip()
    headers = {
        "Cookie": cookie,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36",
        "Referer": "https://bz.feigua.cn/",
    }
    return headers


def load_checkpoint(path: str) -> dict:
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_checkpoint(path: str, data: dict):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    for attempt in range(5):
        try:
            os.replace(tmp, path)
            return
        except PermissionError:
            time.sleep(1)
    # 最后兜底：直接写目标文件
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def ts():
    """当前毫秒时间戳"""
    return int(time.time() * 1000)


def clean_html_tags(text: str) -> str:
    """清除HTML标签"""
    if not text:
        return ""
    return re.sub(r"<[^>]+>", "", text)


def parse_num(val) -> float | str:
    """将带单位的字符串转为数字: '105.8w' -> 1058000, '12.41%' -> '12.41%', '3874' -> 3874"""
    if val is None or val == "":
        return ""
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if not s or s == "0":
        return 0
    # 百分比保持原样
    if s.endswith("%"):
        return s
    # 带w/万后缀
    if s.lower().endswith("w") or s.endswith("万"):
        try:
            return round(float(s[:-1]) * 10000)
        except ValueError:
            return s
    # 带亿后缀
    if s.endswith("亿"):
        try:
            return round(float(s[:-1]) * 100000000)
        except ValueError:
            return s
    # 纯数字
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except ValueError:
        return s


def split_category(tname: str) -> tuple[str, str]:
    """拆分分区: '生活-日常' -> ('生活', '日常')"""
    if not tname:
        return "", ""
    parts = tname.split("-", 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return parts[0], ""


def calc_play_fans_rate(avg_play, fans) -> str:
    """计算播粉比"""
    avg_play_num = parse_num(avg_play)
    fans_num = parse_num(fans)
    if isinstance(avg_play_num, (int, float)) and isinstance(fans_num, (int, float)) and fans_num > 0:
        rate = avg_play_num / fans_num * 100
        return f"{rate:.2f}%"
    return ""


# ============ API 请求 ============

class PermissionLimitError(Exception):
    """会员权限限制，无需重试"""
    pass


def api_get(session: requests.Session, headers: dict, url: str, params: dict) -> dict | None:
    """通用API请求，带重试。遇到会员限制直接抛异常。"""
    params["_"] = ts()
    for attempt in range(5):
        try:
            resp = session.get(url, params=params, headers=headers, timeout=30)
            resp.raise_for_status()
            body = resp.json()
            code = body.get("Code")
            if code in (0, 200):
                return body.get("Data", {})
            elif code == 403:
                raise PermissionLimitError(clean_html_tags(body.get("Msg", "")))
            else:
                wait = 10 * (attempt + 1)
                tqdm.write(f"  API异常 code={code} msg={body.get('Msg')}, 等{wait}s...")
                time.sleep(wait)
        except PermissionLimitError:
            raise
        except Exception as e:
            if attempt < 4:
                wait = 5 * (attempt + 1)
                tqdm.write(f"  请求异常: {e}, 等{wait}s...")
                time.sleep(wait)
            else:
                tqdm.write(f"  5次重试均失败: {e}")
                return None
    return None


# ============ 阶段一: 抓取MCN列表 ============

def _fetch_mcn_by_sort(session: requests.Session, headers: dict,
                       sort_val: int, sort_desc: str) -> list[dict]:
    """按指定排序抓取MCN列表"""
    url = f"{BASE_URL}/v1/BloggerInfo/SearchMcn"
    all_mcn = []
    page = 1
    page_size = 10

    print(f"\n  排序方式: {sort_desc} (sort={sort_val})")

    while True:
        params = {
            "pageSize": page_size,
            "keyWord": "",
            "sort": sort_val,
            "Cate": 0,
            "Fans": 0,
            "BloggerCount": 0,
            "Page": page,
        }
        try:
            data = api_get(session, headers, url, params)
        except PermissionLimitError as e:
            print(f"  已达会员上限: {e}")
            break
        if data is None:
            print(f"  第{page}页请求失败，停止")
            break

        results = data.get("Result", [])
        if not results:
            break

        for item in results:
            name = item.get("McnName", "")
            if name:
                all_mcn.append({
                    "McnName": name,
                    "Fans": item.get("Fans", ""),
                    "BloggerCount": item.get("BloggerCount", 0),
                })

        print(f"  第{page}页: +{len(results)}条, 累计{len(all_mcn)}个")
        page += 1
        time.sleep(REQUEST_DELAY)

    print(f"  {sort_desc}完成: {len(all_mcn)} 个MCN")
    return all_mcn


def phase1_fetch_mcn(session: requests.Session, headers: dict) -> list[str]:
    """抓取所有MCN机构名称（sort=0按达人数 + sort=1按粉丝量，增量合并）"""
    print("\n" + "=" * 50)
    print("  阶段一: 抓取MCN机构列表")
    print("=" * 50)

    # 加载已有数据
    ckpt = load_checkpoint(MCN_CHECKPOINT)
    existing_mcn = ckpt.get("mcn_detail", [])
    existing_names = set(m["McnName"] for m in existing_mcn)
    completed_sorts = set(ckpt.get("completed_sorts", []))

    if existing_mcn:
        print(f"已有: {len(existing_mcn)} 个MCN")

    sorts = [(0, "按达人数排序"), (1, "按粉丝量排序")]
    new_count = 0

    for sort_val, sort_desc in sorts:
        if sort_val in completed_sorts:
            print(f"\n  {sort_desc} (sort={sort_val}): 已完成，跳过")
            continue

        fetched = _fetch_mcn_by_sort(session, headers, sort_val, sort_desc)
        added = 0
        for item in fetched:
            name = item["McnName"]
            if name not in existing_names:
                existing_names.add(name)
                existing_mcn.append(item)
                added += 1
        new_count += added
        completed_sorts.add(sort_val)
        print(f"  新增: {added} 个MCN")

        # 每种排序完成后保存
        save_checkpoint(MCN_CHECKPOINT, {
            "mcn_names": [m["McnName"] for m in existing_mcn],
            "mcn_detail": existing_mcn,
            "completed_sorts": list(completed_sorts),
            "updated_at": datetime.now().isoformat(),
        })

    mcn_names = [m["McnName"] for m in existing_mcn]
    print(f"\n阶段一完成: 共 {len(mcn_names)} 个MCN（本次新增 {new_count} 个）")
    return mcn_names


def _save_mcn_checkpoint(all_mcn: list):
    mcn_names = [m["McnName"] for m in all_mcn]
    save_checkpoint(MCN_CHECKPOINT, {
        "mcn_names": mcn_names,
        "mcn_detail": all_mcn,
        "updated_at": datetime.now().isoformat(),
    })


# ============ 阶段二: 抓取每个MCN的旗下达人（页级断点） ============

def _save_blogger_checkpoint(blogger_data: dict):
    save_checkpoint(BLOGGER_CHECKPOINT, {
        "blogger_data": blogger_data,
        "updated_at": datetime.now().isoformat(),
    })


def phase2_fetch_bloggers(session: requests.Session, headers: dict,
                          mcn_names: list[str]) -> dict:
    """抓取所有MCN的旗下达人（页级断点）

    断点结构: blogger_data = {
        "MCN名称": {
            "pages": {"1": [...], "2": [...]},
            "total": 679,
            "done": true/false
        }
    }
    """
    print("\n" + "=" * 50)
    print("  阶段二: 抓取MCN旗下达人")
    print("=" * 50)

    ckpt = load_checkpoint(BLOGGER_CHECKPOINT)
    blogger_data = ckpt.get("blogger_data", {})

    # 统计进度
    done_mcn = [n for n in mcn_names if n in blogger_data and blogger_data[n].get("done")]
    partial_mcn = [n for n in mcn_names if n in blogger_data and not blogger_data[n].get("done")]
    todo_mcn = [n for n in mcn_names if n not in blogger_data]

    total_bloggers = sum(
        sum(len(v) for v in m.get("pages", {}).values())
        for m in blogger_data.values()
    )

    print(f"总计: {len(mcn_names)} 个MCN")
    print(f"已完成: {len(done_mcn)}")
    if partial_mcn:
        print(f"部分完成: {len(partial_mcn)} (将续抓)")
    print(f"未开始: {len(todo_mcn)}")
    print(f"已有达人数据: {total_bloggers} 条")

    # 待处理: 部分完成的 + 未开始的
    work_list = partial_mcn + todo_mcn
    if not work_list:
        print("所有MCN旗下达人已抓取完成")
        return blogger_data

    url = f"{BASE_URL}/v1/BloggerInfo/SearchBlogger"
    page_size = 50
    hit_limit = False

    for i, mcn_name in enumerate(work_list):
        if hit_limit:
            break

        # 恢复该MCN的已有页数据
        mcn_entry = blogger_data.get(mcn_name, {"pages": {}, "total": 0, "done": False})
        pages = mcn_entry["pages"]
        fetched_count = sum(len(v) for v in pages.values())
        start_page = max((int(p) for p in pages), default=0) + 1

        if start_page > 1:
            print(f"  [{i+1}/{len(work_list)}] MCN: {mcn_name[:30]}... (续抓,第{start_page}页起)", end="", flush=True)
        else:
            print(f"  [{i+1}/{len(work_list)}] MCN: {mcn_name[:30]}...", end="", flush=True)

        page = start_page
        while True:
            params = {
                "Cate": 0,
                "McnName": mcn_name,
                "Page": page,
                "sort": 7,
                "pageSize": page_size,
            }
            try:
                data = api_get(session, headers, url, params)
            except PermissionLimitError as e:
                print(f"\n  已达会员上限: {e}")
                hit_limit = True
                break

            if data is None:
                break

            items = data.get("list", [])
            if not items:
                break

            pages[str(page)] = items
            fetched_count += len(items)
            total = data.get("total", 0)
            mcn_entry["total"] = total

            if fetched_count >= total or len(items) < page_size:
                mcn_entry["done"] = True
                break

            page += 1
            time.sleep(REQUEST_DELAY)

        # 每个MCN每页都保存
        blogger_data[mcn_name] = mcn_entry
        _save_blogger_checkpoint(blogger_data)

        mcn_bloggers = fetched_count
        total_bloggers = sum(
            sum(len(v) for v in m.get("pages", {}).values())
            for m in blogger_data.values()
        )
        done_count = sum(1 for m in blogger_data.values() if m.get("done"))
        status = "完成" if mcn_entry.get("done") else "中断"
        print(f" -> {mcn_bloggers}人[{status}] (累计{total_bloggers}人, {done_count}MCN完成)")

    _save_blogger_checkpoint(blogger_data)

    total = sum(
        sum(len(v) for v in m.get("pages", {}).values())
        for m in blogger_data.values()
    )
    done_count = sum(1 for m in blogger_data.values() if m.get("done"))
    print(f"\n阶段二: {done_count}/{len(mcn_names)} 个MCN完成, 共 {total} 条达人数据")
    return blogger_data


# ============ 粉丝榜补漏: API 1 抓榜单 + API 2 补字段 ============

def _get_existing_uids(blogger_data: dict) -> set:
    """从MCN达人数据中收集所有已有的B站UID"""
    uids = set()
    for mcn_entry in blogger_data.values():
        for items in mcn_entry.get("pages", {}).values():
            for b in items:
                uid = b.get("Uid", "")
                if uid:
                    uids.add(str(uid))
    return uids


def phase_rank_fetch(session: requests.Session, headers: dict,
                     blogger_data: dict) -> list[dict]:
    """抓取粉丝榜Top1000，去掉MCN已有的，用API 2补充详情"""
    print("\n" + "=" * 50)
    print("  粉丝榜补漏: 抓取榜单 → 去重 → 补充详情")
    print("=" * 50)

    ckpt = load_checkpoint(RANK_CHECKPOINT)
    rank_data = ckpt.get("rank_data", {})  # {page_str: [items]}
    detail_data = ckpt.get("detail_data", {})  # {blogger_id_str: detail}

    # --- 步骤1: 抓取粉丝榜 ---
    existing_pages = set(rank_data.keys())
    total_pages = 50  # 1000条 / 20条每页
    pages_todo = [p for p in range(1, total_pages + 1) if str(p) not in existing_pages]

    if pages_todo:
        print(f"\n步骤1: 抓取粉丝榜（已有{len(existing_pages)}页，待抓{len(pages_todo)}页）")
        url = f"{BASE_URL}/v1/Rank/GetFansRank"

        for pg in pages_todo:
            print(f"  榜单第{pg}/{total_pages}页...", end=" ", flush=True)
            try:
                data = api_get(session, headers, url, {
                    "pageSize": 20, "Cate": 0, "BloggerFansType": 0, "page": pg,
                })
            except PermissionLimitError as e:
                print(f"\n  已达会员上限: {e}")
                break

            if data is None:
                print("失败")
                continue

            items = data.get("Result", [])
            if not items:
                print("空")
                break

            rank_data[str(pg)] = items
            print(f"{len(items)}条")

            save_checkpoint(RANK_CHECKPOINT, {
                "rank_data": rank_data,
                "detail_data": detail_data,
                "updated_at": datetime.now().isoformat(),
            })
            time.sleep(REQUEST_DELAY)
    else:
        print(f"\n步骤1: 粉丝榜已完成（{len(existing_pages)}页）")

    # --- 步骤2: 去重 ---
    existing_uids = _get_existing_uids(blogger_data)
    rank_bloggers = []  # (Id, UId) - 需要补详情的
    for pg in sorted(rank_data.keys(), key=int):
        for item in rank_data[pg]:
            uid = str(item.get("UId", ""))
            bid = item.get("Id")
            if uid and uid not in existing_uids and bid:
                rank_bloggers.append((bid, uid))

    # 也按BloggerId去重
    seen_bids = set()
    unique_rank = []
    for bid, uid in rank_bloggers:
        if bid not in seen_bids:
            unique_rank.append((bid, uid))
            seen_bids.add(bid)

    print(f"\n步骤2: 粉丝榜共{sum(len(v) for v in rank_data.values())}人，"
          f"MCN已有{len(existing_uids)}人，需补漏{len(unique_rank)}人")

    # --- 步骤3: 用API 2补充详情 ---
    completed_details = set(detail_data.keys())
    detail_todo = [(bid, uid) for bid, uid in unique_rank if str(bid) not in completed_details]

    if detail_todo:
        print(f"\n步骤3: 补充详情（已有{len(completed_details)}条，待抓{len(detail_todo)}条）")
        url = f"{BASE_URL}/V1/BloggerInfo/DetailNew"

        for i, (bid, uid) in enumerate(detail_todo):
            print(f"  [{i+1}/{len(detail_todo)}] BloggerId={bid}...", end=" ", flush=True)
            try:
                data = api_get(session, headers, url, {"bloggerId": bid})
            except PermissionLimitError as e:
                print(f"\n  已达会员上限: {e}")
                break

            if data:
                detail_data[str(bid)] = data
                name = data.get("BloggerInfo", {}).get("BloggerName", "")
                print(f"{name[:20]}")
            else:
                print("失败")

            if (i + 1) % 10 == 0:
                save_checkpoint(RANK_CHECKPOINT, {
                    "rank_data": rank_data,
                    "detail_data": detail_data,
                    "updated_at": datetime.now().isoformat(),
                })
            time.sleep(REQUEST_DELAY)

        save_checkpoint(RANK_CHECKPOINT, {
            "rank_data": rank_data,
            "detail_data": detail_data,
            "updated_at": datetime.now().isoformat(),
        })
    else:
        print(f"\n步骤3: 详情已全部补充完成（{len(completed_details)}条）")

    print(f"\n补漏完成: 共补充 {len(detail_data)} 个独有达人的详情")
    return detail_data


def build_row_from_detail(detail: dict, rank_item: dict = None) -> dict:
    """从API 2详情数据构建Excel行（用于粉丝榜补漏的达人）
    列名与 build_row 完全一致，确保导出格式统一。
    """
    bi = detail.get("BloggerInfo", {})
    bv = detail.get("BloggerVideo", {})
    bl = detail.get("BloggerLive", {})
    cat1, cat2 = split_category(bi.get("TName", ""))

    fans_raw = bi.get("Fans", "")
    avg_play_raw = bv.get("AvgPlayCount", "")

    row = {
        # ---- 基本信息（与build_row一致） ----
        "UP主昵称": bi.get("BloggerName", ""),
        "B站UID": bi.get("MId", ""),
        "飞瓜ID": rank_item.get("Id", "") if rank_item else "",
        "性别": {0: "未知", 1: "男", 2: "女"}.get(bi.get("Sex", 0), "未知"),
        "B站等级": bi.get("LevelNumber", ""),
        "大会员类型": bi.get("VipType", ""),
        "一级分区": cat1,
        "二级分区": cat2,
        "MCN机构": bi.get("McnName", ""),
        "地区": bi.get("RegionName", ""),
        "城市": bi.get("SecondRegionName", ""),
        "IP归属地": bi.get("IpRegionName", ""),
        "认证": "是" if bi.get("OfficialVerified") else "否",
        "认证头衔": bi.get("OfficialTitle", ""),
        "个性签名": bi.get("Sign", ""),
        "飞瓜评分": parse_num(bi.get("Score", "")),
        "标签": ", ".join(bi.get("BloggerVideoTags", []) or []),

        # ---- 粉丝与互动 ----
        "粉丝数": rank_item.get("FanCount", "") if rank_item else parse_num(fans_raw),
        "主要粉丝性别": "",
        "近期投稿数": "",
        "平均播放量": parse_num(avg_play_raw),
        "平均点赞数": parse_num(bv.get("AvgLikeCount", "")),
        "平均收藏数": parse_num(bv.get("AvgCollectCount", "")),
        "平均投币数": parse_num(bv.get("AvgCoinCount", "")),
        "平均评论数": parse_num(bv.get("AvgReplyCount", "")),
        "互动率": bv.get("InteractRate", ""),
        "播粉比": bv.get("PlayFansRate", "") or calc_play_fans_rate(avg_play_raw, fans_raw),

        # ---- 商业报价 ----
        "有广告报价": "",
        "植入报价": "",
        "定制报价": "",
        "植入方式": "",
        "定制方式": "",

        # ---- 带货 ----
        "有带货": "",
        "主要带货品类": "",
        "带货品类明细": "",

        # ---- 状态标记 ----
        "有直播": "是" if bi.get("HasLive") else "否",
        "有联系方式": "",
        "有品牌合作": "",
        "粉丝勋章": "",

        # ---- 数据来源标记 ----
        "数据来源": "粉丝榜补漏",
    }

    return row


# ============ 数据导出 ============

def extract_tags(tag_list: list) -> str:
    """从Tag列表提取标签文本"""
    if not tag_list:
        return ""
    names = []
    for t in tag_list:
        text = clean_html_tags(t.get("TagName", ""))
        if text:
            names.append(text)
    return ", ".join(names)


def extract_product_cates(cate_list: list) -> str:
    """从ProductCateList提取品类"""
    if not cate_list:
        return ""
    return ", ".join(f"{c['CateName']}({c['Count']})" for c in cate_list if c.get("CateName"))


def build_row(item: dict, mcn_name: str) -> dict:
    """将一条MCN达人数据转为Excel行"""
    cat1, cat2 = split_category(item.get("TName", ""))
    fans = parse_num(item.get("Fans", ""))
    avg_play = parse_num(item.get("AvgPlayCount", ""))

    row = {
        # ---- 基本信息 ----
        "UP主昵称": item.get("BloggerName", ""),
        "B站UID": item.get("Uid", ""),
        "飞瓜ID": item.get("BloggerId", ""),
        "性别": {0: "未知", 1: "男", 2: "女"}.get(item.get("Sex", 0), "未知"),
        "B站等级": item.get("LevelNumber", ""),
        "大会员类型": item.get("VipType", ""),
        "一级分区": cat1,
        "二级分区": cat2,
        "MCN机构": mcn_name,
        "地区": item.get("RegionName", ""),
        "城市": item.get("SecondRegionName", ""),
        "IP归属地": item.get("IpRegionName", ""),
        "认证": "是" if item.get("OfficialVerified") else "否",
        "认证头衔": item.get("OfficialTitle", ""),
        "个性签名": item.get("Sign", ""),
        "飞瓜评分": parse_num(item.get("Score", "")),
        "标签": extract_tags(item.get("Tag", [])),

        # ---- 粉丝与互动 ----
        "粉丝数": fans,
        "主要粉丝性别": item.get("MainFans", ""),
        "近期投稿数": parse_num(item.get("ArchiveCount", "")),
        "平均播放量": avg_play,
        "平均点赞数": parse_num(item.get("AvgLikeCount", "")),
        "平均收藏数": parse_num(item.get("AvgCollectCount", "")),
        "平均投币数": parse_num(item.get("AvgCoinCount", "")),
        "平均评论数": parse_num(item.get("AvgReplyCount", "")),
        "互动率": item.get("InteractRatestr", ""),
        "播粉比": calc_play_fans_rate(item.get("AvgPlayCount", ""), item.get("Fans", "")),

        # ---- 商业报价 ----
        "有广告报价": "是" if item.get("HasAdPrice") else "否",
        "植入报价": parse_num(item.get("ImplantPrice", "")),
        "定制报价": parse_num(item.get("CustomPrice", "")),
        "植入方式": item.get("ImplanttypeStr", ""),
        "定制方式": item.get("CustomtypeStr", ""),

        # ---- 带货 ----
        "有带货": "是" if item.get("HasProduct") else "否",
        "主要带货品类": item.get("MainProductCate", "") or "",
        "带货品类明细": extract_product_cates(item.get("ProductCateList", [])),

        # ---- 状态标记 ----
        "有直播": "是" if item.get("HasLive") else "否",
        "有联系方式": "是" if item.get("HasTel") else "否",
        "有品牌合作": "是" if item.get("HasBrand") else "否",
        "粉丝勋章": "是" if item.get("FansBadge") else "否",

        # ---- 数据来源标记 ----
        "数据来源": "MCN达人",
    }

    return row


def save_to_excel(rows: list[dict], filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "飞瓜UP主数据"

    if not rows:
        print("没有数据可保存")
        return

    # 收集所有可能的列（因为有/无详情字段数不同）
    all_headers = []
    seen = set()
    for row in rows:
        for k in row.keys():
            if k not in seen:
                all_headers.append(k)
                seen.add(k)

    for col, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # 清除Excel不允许的控制字符
    ILLEGAL_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

    for row_idx, row_data in enumerate(rows, 2):
        for col, header in enumerate(all_headers, 1):
            val = row_data.get(header, "")
            if isinstance(val, str):
                val = ILLEGAL_RE.sub("", val)
            ws.cell(row=row_idx, column=col, value=val)

    for col_idx, header in enumerate(all_headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(str(header)) * 2 + 4, 12)

    wb.save(filename)
    print(f"数据已保存到: {filename}")


# ============ 主流程 ============

def main():
    print("=" * 50)
    print("  飞瓜平台B站UP主数据抓取")
    print("=" * 50)

    ensure_dir()
    headers = load_config()
    session = requests.Session()

    # 验证Cookie
    print("验证Cookie...", end=" ", flush=True)
    test = api_get(session, headers,
                   f"{BASE_URL}/v1/Rank/GetFansRank",
                   {"pageSize": 1, "Cate": 0, "BloggerFansType": 0, "page": 1})
    if test is None:
        print("失败，请检查Cookie")
        return
    print(f"成功（榜单共{test.get('TotalCount', '?')}条）")

    # 主菜单
    while True:
        print("\n" + "-" * 40)
        print("  请选择操作:")
        print("  1. 完整抓取（MCN列表 → 旗下达人）")
        print("  2. 粉丝榜补漏（抓榜单独有达人 + 补详情）")
        print("  3. 导出Excel（合并所有数据）")
        print("  4. 查看当前进度")
        print("  0. 退出")
        print("-" * 40)

        choice = input("请输入选项: ").strip()

        if choice == "1":
            mcn_names = phase1_fetch_mcn(session, headers)
            if mcn_names:
                phase2_fetch_bloggers(session, headers, mcn_names)

        elif choice == "2":
            ckpt = load_checkpoint(BLOGGER_CHECKPOINT)
            blogger_data = ckpt.get("blogger_data", {})
            if not blogger_data:
                print("\n错误: 请先执行步骤1")
                continue
            phase_rank_fetch(session, headers, blogger_data)

        elif choice == "3":
            # --- 导出Excel: 合并MCN达人 + 粉丝榜补漏 ---
            ckpt = load_checkpoint(BLOGGER_CHECKPOINT)
            blogger_data = ckpt.get("blogger_data", {})

            rank_ckpt = load_checkpoint(RANK_CHECKPOINT)
            rank_data = rank_ckpt.get("rank_data", {})
            rank_detail = rank_ckpt.get("detail_data", {})

            # Part 1: MCN达人数据
            all_rows = []
            seen_uids = set()

            if blogger_data:
                for mcn_name, mcn_entry in blogger_data.items():
                    pages = mcn_entry.get("pages", {})
                    for pg in sorted(pages.keys(), key=int):
                        for item in pages[pg]:
                            uid = str(item.get("Uid", ""))
                            if uid and uid in seen_uids:
                                continue
                            seen_uids.add(uid)
                            all_rows.append(build_row(item, mcn_name))

            mcn_count = len(all_rows)

            # Part 2: 粉丝榜补漏达人
            rank_added = 0
            if rank_data:
                # 建立 Id -> rank_item 映射
                rank_items_by_id = {}
                for pg in rank_data.values():
                    for item in pg:
                        rank_items_by_id[str(item.get("Id", ""))] = item

                for bid_str, detail in rank_detail.items():
                    uid = str(detail.get("BloggerInfo", {}).get("MId", ""))
                    if uid and uid not in seen_uids:
                        seen_uids.add(uid)
                        rank_item = rank_items_by_id.get(bid_str, {})
                        all_rows.append(build_row_from_detail(detail, rank_item))
                        rank_added += 1

            if not all_rows:
                print("\n错误: 没有任何数据，请先执行抓取")
                continue

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"feigua_blogger_{timestamp}.xlsx"
            save_to_excel(all_rows, filename)

            print(f"\n总计: {len(all_rows)} 条")
            print(f"  MCN达人: {mcn_count} 条")
            if rank_added:
                print(f"  粉丝榜补漏: {rank_added} 条")
            print(f"文件: {filename}")

        elif choice == "4":
            mcn_ckpt = load_checkpoint(MCN_CHECKPOINT)
            blog_ckpt = load_checkpoint(BLOGGER_CHECKPOINT)
            rank_ckpt = load_checkpoint(RANK_CHECKPOINT)

            mcn_count = len(mcn_ckpt.get("mcn_names", []))
            blog_data = blog_ckpt.get("blogger_data", {})
            mcn_total = len(blog_data)
            mcn_done = sum(1 for m in blog_data.values() if m.get("done"))
            mcn_partial = mcn_total - mcn_done
            blogger_total = sum(
                sum(len(v) for v in m.get("pages", {}).values())
                for m in blog_data.values()
            )
            rank_pages = len(rank_ckpt.get("rank_data", {}))
            rank_details = len(rank_ckpt.get("detail_data", {}))

            print(f"\n  MCN机构: {mcn_count} 个")
            print(f"  MCN达人: {mcn_done} 完成, {mcn_partial} 部分完成, 共 {blogger_total} 条")
            print(f"  粉丝榜: {rank_pages}/50 页")
            print(f"  粉丝榜补漏详情: {rank_details} 条")

        elif choice == "0":
            print("退出")
            break
        else:
            print("无效选项")


if __name__ == "__main__":
    main()
