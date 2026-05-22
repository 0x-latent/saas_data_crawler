"""
合并飞瓜和火花两份数据表
以飞瓜表为基准，通过 B站UID = UP主MID 匹配，将火花独有字段追加到右侧

用法: .venv/Scripts/python.exe merge_data.py
"""

import os
import re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([os.path.join(".venv", "Scripts", "pip.exe"), "install", "openpyxl"])
    import openpyxl

# ============ 配置 ============
FEIGUA_FILE = "data/feigua_blogger_20260519_161712.xlsx"
HUOHUA_FILE = "data/huohua_full_20260430_145400.xlsx"

FEIGUA_UID_COL = "B站UID"    # 飞瓜表中的UID列名
HUOHUA_MID_COL = "UP主MID"   # 火花表中的MID列名


def normalize_uid(val) -> str:
    """统一UID格式: int/float/str 都转为纯数字字符串"""
    if val is None:
        return ""
    if isinstance(val, float):
        return str(int(val))
    s = str(val).strip()
    # 去除可能的 .0 后缀
    if s.endswith(".0"):
        s = s[:-2]
    return s


def read_xlsx(path: str) -> tuple[list[str], list[dict]]:
    """读取Excel，返回 (列名列表, 行数据列表)"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.rows)
    headers = [cell.value for cell in rows[0]]

    data = []
    for row in rows[1:]:
        values = [cell.value for cell in row]
        data.append(dict(zip(headers, values)))

    wb.close()
    return headers, data


def save_xlsx(headers: list[str], rows: list[dict], filename: str):
    """保存到Excel"""
    ILLEGAL_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合并数据"

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    for row_idx, row_data in enumerate(rows, 2):
        for col, header in enumerate(headers, 1):
            val = row_data.get(header, "")
            if isinstance(val, str):
                val = ILLEGAL_RE.sub("", val)
            ws.cell(row=row_idx, column=col, value=val)

    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(str(header)) * 2 + 4, 12)

    wb.save(filename)
    print(f"已保存: {filename}")


def main():
    print("=" * 50)
    print("  合并飞瓜 + 火花数据")
    print("=" * 50)

    # 读取两份数据
    print(f"\n读取飞瓜表: {FEIGUA_FILE}")
    fg_headers, fg_data = read_xlsx(FEIGUA_FILE)
    print(f"  {len(fg_data)} 行, {len(fg_headers)} 列")

    print(f"\n读取火花表: {HUOHUA_FILE}")
    hh_headers, hh_data = read_xlsx(HUOHUA_FILE)
    print(f"  {len(hh_data)} 行, {len(hh_headers)} 列")

    # 火花表按MID建索引（统一格式）
    hh_by_mid = {}
    for row in hh_data:
        mid = normalize_uid(row.get(HUOHUA_MID_COL))
        if mid:
            hh_by_mid[mid] = row
    print(f"\n火花表有效MID: {len(hh_by_mid)} 个")

    # 找出火花独有列（不在飞瓜表中的），排除匹配键
    fg_col_set = set(fg_headers)
    hh_only_cols = [h for h in hh_headers if h not in fg_col_set and h != HUOHUA_MID_COL]
    print(f"火花独有列（将追加）: {len(hh_only_cols)} 个")

    # 合并: 飞瓜列 + 火花独有列
    merged_headers = fg_headers + hh_only_cols

    matched = 0
    price_filled = 0
    for row in fg_data:
        uid = normalize_uid(row.get(FEIGUA_UID_COL))
        if uid:
            hh_row = hh_by_mid.get(uid)
            if hh_row:
                matched += 1
                # 追加火花独有列
                for col in hh_only_cols:
                    row[col] = hh_row.get(col, "")
                # 飞瓜缺报价时，用火花报价补充
                fg_implant = row.get("植入报价")
                fg_custom = row.get("定制报价")
                if not fg_implant or fg_implant == "" or fg_implant == 0:
                    hh_val = hh_row.get("植入视频报价", "")
                    if hh_val:
                        row["植入报价"] = hh_val
                        price_filled += 1
                if not fg_custom or fg_custom == "" or fg_custom == 0:
                    hh_val = hh_row.get("定制视频报价", "")
                    if hh_val:
                        row["定制报价"] = hh_val
                        price_filled += 1

    print(f"\n匹配结果: {matched}/{len(fg_data)} 条飞瓜数据匹配到火花数据")
    print(f"报价补充: {price_filled} 次（飞瓜缺报价时用火花数据填充）")
    print(f"合并后: {len(fg_data)} 行, {len(merged_headers)} 列 ({len(fg_headers)}飞瓜 + {len(hh_only_cols)}火花)")

    # 保存
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"data/merged_{timestamp}.xlsx"
    save_xlsx(merged_headers, fg_data, filename)

    print(f"\n完成！")
    print(f"  飞瓜原始: {len(fg_headers)} 列")
    print(f"  火花追加: {len(hh_only_cols)} 列")
    print(f"  匹配率: {matched}/{len(fg_data)} ({matched/len(fg_data)*100:.1f}%)")
    print(f"  文件: {filename}")


if __name__ == "__main__":
    main()
