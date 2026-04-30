"""
B站火花平台UP主数据抓取脚本
多Cookie并发抓取，支持断点续抓
进度保存在 checkpoint.json，中断后重新运行自动续抓

用法:
  1. 在 config.yaml 的 cookies 列表中填入一个或多个 Cookie
  2. 运行: .venv/Scripts/python.exe scraper.py
  3. 中断后换/加 Cookie，重新运行即可续抓
"""

import requests
import time
import json
import os
import yaml
from datetime import datetime
from tqdm import tqdm
from concurrent.futures import ThreadPoolExecutor
import threading
import queue

try:
    import openpyxl
except ImportError:
    print("正在安装 openpyxl...")
    import subprocess
    subprocess.check_call([os.path.join(".venv", "Scripts", "pip.exe"), "install", "openpyxl"])
    import openpyxl


# ============ 配置区 ============
CHECKPOINT_FILE = "checkpoint.json"
CONFIG_FILE = "config.yaml"

BASE_URL = "https://huahuo.bilibili.com/commercialorder/api/web_api/v1/advertiser/upper_square/search"
CT_ID = "XBhqj6L6Amz9L13roEtHN"

PAGE_SIZE = 20
# 同一个Cookie的请求间隔（秒）
REQUEST_DELAY = 2.0
SAVE_EVERY = 20


def load_config() -> list[dict]:
    """加载配置，返回 headers 列表（每个Cookie一套）"""
    if not os.path.exists(CONFIG_FILE):
        print(f"错误: 找不到 {CONFIG_FILE}，请创建并填入 cookies")
        exit(1)
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    cookies = cfg.get("cookies", [])
    # 兼容旧版单cookie格式
    if not cookies and cfg.get("cookie"):
        cookies = [cfg["cookie"]]

    if not cookies:
        print(f"错误: {CONFIG_FILE} 中未配置任何 cookie")
        exit(1)

    headers_list = []
    for cookie in cookies:
        cookie = cookie.strip()
        if cookie:
            headers_list.append({
                "Cookie": cookie,
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36",
                "Referer": "https://huahuo.bilibili.com/",
            })

    return headers_list


# ============ 断点管理 ============

def load_checkpoint() -> dict:
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_checkpoint(page_results: dict, total_pages: int):
    data = {
        "total_pages": total_pages,
        "completed_pages": sorted(page_results.keys()),
        "page_data": page_results,
        "updated_at": datetime.now().isoformat(),
    }
    tmp = CHECKPOINT_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, CHECKPOINT_FILE)


def clear_checkpoint():
    if os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)


# ============ 请求 ============

def fetch_page(page: int, session: requests.Session, headers: dict) -> list | None:
    """抓取单页，返回数据列表。触发限频时自动等待重试，最多5次。"""
    params = {"ct_id": CT_ID, "page": page}
    for attempt in range(5):
        try:
            resp = session.get(BASE_URL, params=params, headers=headers, timeout=30)
            resp.raise_for_status()
            body = resp.json()

            if body.get("code") == 1031 or body.get("status") == "fail":
                wait = 10 * (attempt + 1)
                tqdm.write(f"  [页{page}] 频率限制，等{wait}s... (第{attempt+1}次)")
                time.sleep(wait)
                continue

            result = body["result"]
            if isinstance(result, str):
                wait = 10 * (attempt + 1)
                tqdm.write(f"  [页{page}] 异常响应: {body.get('msg', result)}, 等{wait}s...")
                time.sleep(wait)
                continue

            return result.get("data", [])
        except Exception as e:
            if attempt < 4:
                wait = 5 * (attempt + 1)
                tqdm.write(f"  [页{page}] 请求异常: {e}, 等{wait}s...")
                time.sleep(wait)
            else:
                tqdm.write(f"  [页{page}] 5次重试均失败: {e}")
                return None


# ============ Worker（每个Cookie一个线程） ============

def worker(worker_id: int, task_queue: queue.Queue, headers: dict,
           page_results: dict, failed_pages: list,
           lock: threading.Lock, pbar: tqdm, total_pages: int):
    """单个Cookie的工作线程，从共享队列取任务，谁快谁多干"""
    session = requests.Session()
    count = 0
    while True:
        try:
            page = task_queue.get_nowait()
        except queue.Empty:
            break

        # 同一Cookie两次请求之间保持间隔
        if count > 0:
            time.sleep(REQUEST_DELAY)

        data = fetch_page(page, session, headers)
        count += 1

        with lock:
            if data is not None:
                page_results[page] = data
            else:
                failed_pages.append(page)
            pbar.update(1)
            pbar.set_postfix(成功=len(page_results), 失败=len(failed_pages))

            if len(page_results) % SAVE_EVERY == 0:
                save_checkpoint(page_results, total_pages)

        task_queue.task_done()

    tqdm.write(f"  [Cookie#{worker_id+1}] 完成 {count} 页")


# ============ 数据提取 ============

def extract_price_info(price_infos: list) -> dict:
    result = {"植入视频报价": "", "定制视频报价": "", "直发动态报价": "", "转发动态报价": ""}
    type_map = {1: "植入视频报价", 2: "定制视频报价", 3: "直发动态报价", 4: "转发动态报价"}
    for p in (price_infos or []):
        key = type_map.get(p.get("cooperation_type"))
        if key:
            result[key] = p.get("platform_price", "")
    return result


def extract_row(item: dict) -> dict:
    prices = extract_price_info(item.get("price_infos", []))
    portrait = item.get("upper_user_portrait_info") or {}
    draft_data = item.get("up_draft_data_view") or {}

    goods_data = item.get("goods_up_data_view") or {}

    return {
        # ---- 基本信息 ----
        "UP主昵称": item.get("nickname", ""),
        "UP主MID": item.get("upper_mid", ""),
        "性别": item.get("gender_desc", ""),
        "地区": item.get("region_desc", ""),
        "城市": item.get("second_region_desc", ""),
        "一级分区": item.get("partition_name", ""),
        "二级分区": item.get("second_partition_name", ""),
        "MCN公司": item.get("mcn_company_name", ""),
        "是否高潜UP主": item.get("is_high_potential", 0),
        "UP主类型": item.get("upper_type", ""),
        "活跃状态": item.get("active_status", ""),
        "是否新UP主": item.get("is_new_upper", 0),
        "是否新入驻": item.get("is_new_enter", 0),
        "是否直播中": item.get("is_live", ""),
        # ---- 核心指标 ----
        "粉丝量": item.get("fans_num", 0),
        "涨粉量": item.get("fans_inc", 0),
        "涨粉率(%)": item.get("fans_inc_rate", 0),
        "播放量中位数": item.get("average_play_cnt", 0),
        "30天播放量中位数": draft_data.get("play_median_nature_30d", 0),
        "互动率": item.get("interactive_rate", 0),
        "热门稿件率": item.get("hot_avid_rate", 0),
        "投稿时长": item.get("draft_duration", 0),
        "近30天投稿数": item.get("avid_cnt_30d", 0),
        "近90天投稿数": item.get("avid_cnt_90d", 0),
        "近180天投稿数": item.get("avid_cnt_180d", 0),
        "180天热门稿件数": item.get("hot_avid_cnt_180d", 0),
        "商业视频数": item.get("commercial_avid_cnt", 0),
        "移动端播放占比": item.get("app_vv_percent", 0),
        "PC端播放占比": item.get("pc_vv_percent", 0),
        "移动端播放占比说明": item.get("app_vv_percent_desc", ""),
        "PC端播放占比说明": item.get("pc_vv_percent_desc", ""),
        "综合评分": item.get("synthetical_score", 0),
        "磁力值": item.get("magnetic_value", ""),
        "动态评分": item.get("dynamic_score", ""),
        "预估播放量": item.get("estimated_play", 0),
        "预估花费": item.get("estimated_cost", ""),
        "是否助推中": item.get("is_boosting", 0),
        # ---- 报价 ----
        "植入视频报价": prices["植入视频报价"],
        "定制视频报价": prices["定制视频报价"],
        "直发动态报价": prices["直发动态报价"],
        "转发动态报价": prices["转发动态报价"],
        "品牌植入CPM": item.get("brand_embedding_platform_price_cpm", ""),
        "内容定制CPM": item.get("content_customized_platform_price_cpm", ""),
        "CPM": item.get("cpm", ""),
        "CPC": item.get("cpc", ""),
        "CPE": item.get("cpe", ""),
        "蓝链点击率": item.get("blue_url_ctr", ""),
        "蓝链点击成本": item.get("blue_url_click_cost", ""),
        # ---- 效果数据(全部流量) ----
        "全部流量_播放中位数_7天": draft_data.get("play_median_all_7d", ""),
        "全部流量_播放中位数_15天": draft_data.get("play_median_all_15d", ""),
        "全部流量_播放中位数_30天": draft_data.get("play_median_all_30d", ""),
        "全部流量_CPM_7天": draft_data.get("cpm_all_7d", ""),
        "全部流量_CPM_15天": draft_data.get("cpm_all_15d", ""),
        "全部流量_CPM_30天": draft_data.get("cpm_all_30d", ""),
        "全部流量_CPC_7天": draft_data.get("cpc_all_7d", ""),
        "全部流量_CPC_15天": draft_data.get("cpc_all_15d", ""),
        "全部流量_CPC_30天": draft_data.get("cpc_all_30d", ""),
        "全部流量_CPE_7天": draft_data.get("cpe_all_7d", ""),
        "全部流量_CPE_30天": draft_data.get("cpe_all_30d", ""),
        # ---- 效果数据(自然流量) ----
        "自然流量_播放中位数_7天": draft_data.get("play_median_nature_7d", ""),
        "自然流量_播放中位数_15天": draft_data.get("play_median_nature_15d", ""),
        "自然流量_播放中位数_30天": draft_data.get("play_median_nature_30d", ""),
        "自然流量_CPM_7天": draft_data.get("cpm_nature_7d", ""),
        "自然流量_CPM_15天": draft_data.get("cpm_nature_15d", ""),
        "自然流量_CPM_30天": draft_data.get("cpm_nature_30d", ""),
        "自然流量_CPC_7天": draft_data.get("cpc_nature_7d", ""),
        "自然流量_CPC_15天": draft_data.get("cpc_nature_15d", ""),
        "自然流量_CPC_30天": draft_data.get("cpc_nature_30d", ""),
        "自然流量_CPE_7天": draft_data.get("cpe_nature_7d", ""),
        "自然流量_CPE_30天": draft_data.get("cpe_nature_30d", ""),
        # ---- 关注用户画像 ----
        "关注用户_男性占比(%)": portrait.get("sax_distributions_1", ""),
        "关注用户_女性占比(%)": portrait.get("sax_distributions_0", ""),
        "关注用户_18-24岁占比(%)": portrait.get("age_distributions_18_24", ""),
        "关注用户_25-30岁占比(%)": portrait.get("age_distributions_25_30", ""),
        "关注用户_30+岁占比(%)": portrait.get("age_distributions_30_999", ""),
        "关注用户_0-17岁占比(%)": portrait.get("age_distributions_0_17", ""),
        # ---- 观看用户画像 ----
        "观看用户_粉丝观看占比(%)": portrait.get("fans_vv_distributions_audience", ""),
        "观看用户_0-17岁占比(%)": portrait.get("age_distributions_audience_0_17", ""),
        "观看用户_18-24岁占比(%)": portrait.get("age_distributions_audience_18_24", ""),
        "观看用户_25-30岁占比(%)": portrait.get("age_distributions_audience_25_30", ""),
        "观看用户_30+岁占比(%)": portrait.get("age_distributions_audience_30_999", ""),
        # ---- DMP人群 ----
        "DMP粉丝匹配率": item.get("dmp_fans_rate", ""),
        "DMP覆盖率": item.get("dmp_cover_rate", ""),
        "DMP覆盖人数": item.get("dmp_cover_count", ""),
        # ---- 直播数据 ----
        "直播_场均观看人数_30天": item.get("live_audience_h_avg_30d", ""),
        "直播_场均弹幕数_30天": item.get("live_danmu_h_avg_30d", ""),
        "直播_人气值": item.get("popularity", ""),
        "直播_等级": item.get("live_level", ""),
        "直播_GMV均值说明": item.get("live_gmv_avg_desc", ""),
        "直播_人均观看时长": item.get("live_per_viewing_time", ""),
        "直播_带货均价": item.get("live_good_price_avg", ""),
        "直播_90天带货金额": item.get("live_sale_amount_90day", ""),
        "直播_带货品类": ", ".join(item.get("live_good_category", []) or []),
        "带货权限": item.get("goods_permission", ""),
        # ---- 带货数据 ----
        "带货等级": goods_data.get("goods_up_level", ""),
        "带货品类": ", ".join(goods_data.get("goods_category", []) or []),
        "带货均价": goods_data.get("goods_avg_price", ""),
        "带货均价说明": goods_data.get("goods_avg_price_desc", ""),
        "带货GPM": goods_data.get("goods_gpm", ""),
        "带货GPM说明": goods_data.get("goods_gpm_desc", ""),
        "全部点击率": goods_data.get("all_click_rate", ""),
        "全部点击率说明": goods_data.get("all_click_rate_desc", ""),
        "评论区点击率": goods_data.get("comment_all_click_rate", ""),
        "评论区点击率说明": goods_data.get("comment_all_click_rate_desc", ""),
        # ---- 标签 ----
        "特殊标签": ", ".join(item.get("special_tags", []) or []),
        "合作模式标签": ", ".join(item.get("cooperation_mode_tags", []) or []),
        "高转化标签": ", ".join(item.get("high_convert_tags", []) or []),
        # ---- 链接 ----
        "头像URL": item.get("head_img", ""),
        "B站主页": f"https://space.bilibili.com/{item.get('upper_mid', '')}",
    }


# ============ 导出 ============

def save_to_excel(rows: list[dict], filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UP主数据"

    if not rows:
        print("没有数据可保存")
        return

    headers = list(rows[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    for row_idx, row_data in enumerate(rows, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col, value=row_data.get(header, ""))

    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(header) * 2 + 4, 12)

    wb.save(filename)
    print(f"数据已保存到: {filename}")


# ============ 主流程 ============

def main():
    print("=" * 50)
    print("B站火花平台UP主数据抓取")
    print("=" * 50)

    headers_list = load_config()
    num_cookies = len(headers_list)
    print(f"已加载 {num_cookies} 个 Cookie")

    if num_cookies == 1:
        print("模式: 单Cookie串行")
    else:
        print(f"模式: {num_cookies} Cookie 并发（每个Cookie独立限速 {REQUEST_DELAY}s/请求）")

    print(f"预估速度: ~{num_cookies / REQUEST_DELAY:.1f} 页/秒")

    # 检查断点
    checkpoint = load_checkpoint()
    page_results = {}
    total_pages = 0

    if checkpoint and checkpoint.get("page_data"):
        total_pages = checkpoint["total_pages"]
        for p_str, p_data in checkpoint["page_data"].items():
            page_results[int(p_str)] = p_data
        completed = set(page_results.keys())
        pages_to_fetch = [p for p in range(1, total_pages + 1) if p not in completed]
        print(f"\n从断点恢复: 已完成 {len(completed)}/{total_pages} 页")
        print(f"本次需抓取: {len(pages_to_fetch)} 页")
    else:
        # 全新抓取，用第一个Cookie获取总数
        print("\n正在获取数据总量...")
        first_data = fetch_page(1, requests.Session(), headers_list[0])
        if first_data is None:
            print("错误: 无法获取第一页数据，请检查 config.yaml 中的 Cookie 是否有效")
            return

        # 获取 total_count
        for attempt in range(5):
            try:
                r = requests.get(BASE_URL, params={"ct_id": CT_ID, "page": 1},
                                 headers=headers_list[0], timeout=30)
                body = r.json()
                if isinstance(body.get("result"), dict):
                    total_count = body["result"]["total_count"]
                    total_pages = (total_count + PAGE_SIZE - 1) // PAGE_SIZE
                    break
                time.sleep(10)
            except Exception:
                time.sleep(10)
        else:
            total_pages = 500
            print(f"  无法获取精确总数，按 {total_pages} 页尝试")

        print(f"共 {total_pages} 页 (每页{PAGE_SIZE}条)")
        page_results[1] = first_data
        pages_to_fetch = list(range(2, total_pages + 1))

    if not pages_to_fetch:
        print("所有页已抓取完成，直接导出...")
    else:
        est_time = len(pages_to_fetch) * REQUEST_DELAY / num_cookies / 60
        print(f"预计耗时: ~{est_time:.1f} 分钟\n")

    failed_pages = []
    lock = threading.Lock()

    if pages_to_fetch:
        pbar = tqdm(total=len(pages_to_fetch), desc="抓取进度", unit="页",
                    bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]")

        # 共享任务队列，谁快谁多干
        task_queue = queue.Queue()
        for page in pages_to_fetch:
            task_queue.put(page)

        with ThreadPoolExecutor(max_workers=num_cookies) as executor:
            futures = []
            for wid, hdrs in enumerate(headers_list):
                futures.append(
                    executor.submit(worker, wid, task_queue, hdrs,
                                    page_results, failed_pages, lock, pbar, total_pages)
                )
            for f in futures:
                f.result()

        pbar.close()

    # 最终保存断点
    save_checkpoint(page_results, total_pages)

    # 组装数据
    all_rows = []
    for page in sorted(page_results.keys()):
        for item in page_results[page]:
            all_rows.append(extract_row(item))

    print(f"\n抓取完成！成功: {len(all_rows)} 条, 失败页: {len(failed_pages)}")
    if failed_pages:
        print(f"失败页码: {sorted(failed_pages)[:20]}{'...' if len(failed_pages) > 20 else ''}")
        print("重新运行脚本可自动续抓失败的页")

    # 保存Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"huohua_upzhu_{timestamp}.xlsx"
    save_to_excel(all_rows, filename)

    if not failed_pages:
        # 保留原始数据用于重新导出，不再自动删除
        print("全部抓取成功，原始数据保留在 checkpoint.json（可用于重新导出）")
        print("如需清除请手动删除: del checkpoint.json")

    print(f"\n文件: {filename}")


if __name__ == "__main__":
    main()
