"""
B站火花平台UP主数据抓取脚本（浏览器版）
通过 Playwright 连接已登录的 Chrome，模拟真实浏览行为
触发验证码时自动弹出浏览器窗口等待手动验证

用法:
  1. 关闭所有 Chrome 窗口
  2. 用调试模式启动 Chrome:
     "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\\Users\\pc\\AppData\\Local\\Google\\Chrome\\User Data"
  3. 在 Chrome 中确认已登录火花平台
  4. 运行: .venv\\Scripts\\python.exe scraper_browser.py

阶段一: 抓取UP主列表（500页 x 20条 = 10000个UP主）
阶段二: 逐个访问UP主详情页，拦截API响应获取完整数据
"""

import asyncio
import json
import os
import time
from datetime import datetime

from playwright.async_api import async_playwright

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([os.path.join(".venv", "Scripts", "pip.exe"), "install", "openpyxl"])
    import openpyxl

from tqdm import tqdm


# ============ 配置 ============
CHECKPOINT_DIR = "checkpoints"
PHASE1_CHECKPOINT = os.path.join(CHECKPOINT_DIR, "phase1_list.json")
PHASE2_CHECKPOINT = os.path.join(CHECKPOINT_DIR, "phase2_detail.json")

CDP_URL = "http://localhost:9222"

# 列表页API
SEARCH_URL = "https://huahuo.bilibili.com/commercialorder/api/web_api/v1/advertiser/upper_square/search"
SEARCH_CT_ID = "XBhqj6L6Amz9L13roEtHN"

# UP主详情页URL模板（用mapping_id）
DETAIL_PAGE_URL = "https://huahuo.bilibili.com/#/upper/page/{mapping_id}?referer=UpperContent"

# 需要拦截的API路径
INTERCEPT_APIS = {
    "/advertiser/portrait": "portrait",
    "/portrait/draft/trend_extra": "draft_trend",
    "/attention_user/trend_extra": "fans_trend",
    "/representative/list": "representative",
}

PAGE_SIZE = 20
# 每N个UP主保存一次断点
SAVE_EVERY = 10
# 验证码检测超时（秒）
CAPTCHA_TIMEOUT = 300


# ============ 工具函数 ============

def ensure_dir():
    os.makedirs(CHECKPOINT_DIR, exist_ok=True)


def load_checkpoint(path: str) -> dict:
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_checkpoint_file(path: str, data: dict):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, path)



def print_banner(text: str):
    print(f"\n{'='*50}")
    print(f"  {text}")
    print(f"{'='*50}\n")


# ============ 阶段一：抓取UP主列表 ============

async def fetch_api_via_browser(page, url: str) -> dict | None:
    """通过浏览器访问API URL，获取JSON响应"""
    try:
        resp = await page.goto(url, wait_until="networkidle", timeout=30000)
        if resp is None:
            return None
        content = await page.content()
        # 页面内容是JSON，被浏览器包裹在<pre>标签中
        text = await page.evaluate("() => document.body.innerText")
        return json.loads(text)
    except Exception as e:
        tqdm.write(f"  请求异常: {e}")
        return None


async def check_and_handle_captcha(page, context) -> bool:
    """检测是否触发验证码，如果是则等待用户手动验证"""
    try:
        text = await page.evaluate("() => document.body.innerText")
        body = json.loads(text)
        if body.get("code") == 1031 or body.get("status") == "fail":
            return True
    except Exception:
        pass

    # 检查页面是否有验证码元素
    captcha = await page.query_selector(".geetest_panel, .captcha-container, #gc-box")
    if captcha:
        return True

    return False


async def wait_for_captcha_resolve(page, context):
    """弹出浏览器窗口等待用户手动验证"""
    print("\n" + "!" * 50)
    print("  触发验证码！请在弹出的浏览器窗口中完成验证")
    print("  验证完成后脚本会自动继续")
    print("!" * 50)

    # 导航到火花平台首页触发验证码弹窗
    await page.goto("https://huahuo.bilibili.com/#/upper/index/content", wait_until="networkidle", timeout=60000)

    # 等待验证码消失（用户手动完成）
    start = time.time()
    while time.time() - start < CAPTCHA_TIMEOUT:
        await asyncio.sleep(2)
        # 检查是否还有验证码
        captcha = await page.query_selector(".geetest_panel, .captcha-container, #gc-box")
        if not captcha:
            # 尝试一个API请求验证是否恢复
            await asyncio.sleep(1)
            test_url = f"{SEARCH_URL}?ct_id={SEARCH_CT_ID}&page=1"
            resp = await page.goto(test_url, wait_until="networkidle", timeout=15000)
            try:
                text = await page.evaluate("() => document.body.innerText")
                body = json.loads(text)
                if body.get("code") == 0:
                    print("  验证通过！继续抓取...\n")
                    return True
            except Exception:
                pass

        # 每30秒提醒一次
        elapsed = int(time.time() - start)
        if elapsed % 30 == 0 and elapsed > 0:
            print(f"  等待验证中... ({elapsed}s/{CAPTCHA_TIMEOUT}s)")

    print("  验证超时，请重新运行脚本")
    return False


async def phase1_fetch_list(context):
    """阶段一：抓取UP主列表"""
    print_banner("阶段一：抓取UP主列表")

    # 加载断点
    ckpt = load_checkpoint(PHASE1_CHECKPOINT)
    page_results = {}
    total_pages = 0

    if ckpt and ckpt.get("page_data"):
        total_pages = ckpt["total_pages"]
        page_results = {int(k): v for k, v in ckpt["page_data"].items()}
        completed = set(page_results.keys())
        pages_to_fetch = [p for p in range(1, total_pages + 1) if p not in completed]
        print(f"从断点恢复: 已完成 {len(completed)}/{total_pages} 页")
        print(f"本次需抓取: {len(pages_to_fetch)} 页")
    else:
        pages_to_fetch = None  # 需要先获取总数

    page = await context.new_page()

    # 如果没有断点，先获取总数
    if pages_to_fetch is None:
        print("正在获取数据总量...")
        url = f"{SEARCH_URL}?ct_id={SEARCH_CT_ID}&page=1"
        body = await fetch_api_via_browser(page, url)

        if body is None or body.get("code") != 0:
            if body and await check_and_handle_captcha(page, context):
                resolved = await wait_for_captcha_resolve(page, context)
                if not resolved:
                    await page.close()
                    return {}
                body = await fetch_api_via_browser(page, url)

        if body is None or body.get("code") != 0:
            print("错误: 无法获取数据，请检查登录状态")
            await page.close()
            return {}

        result = body["result"]
        total_count = result["total_count"]
        total_pages = (total_count + PAGE_SIZE - 1) // PAGE_SIZE
        print(f"共 {total_count} 条数据，{total_pages} 页")

        page_results[1] = result["data"]
        pages_to_fetch = list(range(2, total_pages + 1))

    if not pages_to_fetch:
        print("所有页已抓取完成")
        await page.close()
        # 汇总upper_mid列表
        all_mids = []
        for p in sorted(page_results.keys()):
            for item in page_results[p]:
                all_mids.append(item.get("upper_mid"))
        return {"page_results": page_results, "total_pages": total_pages, "all_mids": all_mids}

    print(f"共 {len(pages_to_fetch)} 页待抓取\n")

    pbar = tqdm(total=len(pages_to_fetch), desc="阶段一-列表", unit="页",
                bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")

    for i, pg in enumerate(pages_to_fetch):
        url = f"{SEARCH_URL}?ct_id={SEARCH_CT_ID}&page={pg}"
        body = await fetch_api_via_browser(page, url)

        # 检查限流
        if body and (body.get("code") == 1031 or body.get("status") == "fail"):
            tqdm.write(f"  [页{pg}] 触发限流，等待验证...")
            # 先保存断点
            save_checkpoint_file(PHASE1_CHECKPOINT, {
                "total_pages": total_pages,
                "page_data": {str(k): v for k, v in page_results.items()},
            })
            resolved = await wait_for_captcha_resolve(page, context)
            if not resolved:
                break
            # 重试当前页
            body = await fetch_api_via_browser(page, url)

        if body and body.get("code") == 0 and isinstance(body.get("result"), dict):
            data = body["result"].get("data", [])
            if data:
                page_results[pg] = data
        else:
            tqdm.write(f"  [页{pg}] 获取失败，跳过")

        pbar.update(1)
        pbar.set_postfix(成功=len(page_results))

        # 定期保存断点
        if (i + 1) % SAVE_EVERY == 0:
            save_checkpoint_file(PHASE1_CHECKPOINT, {
                "total_pages": total_pages,
                "page_data": {str(k): v for k, v in page_results.items()},
            })


    pbar.close()
    await page.close()

    # 最终保存
    save_checkpoint_file(PHASE1_CHECKPOINT, {
        "total_pages": total_pages,
        "page_data": {str(k): v for k, v in page_results.items()},
    })

    # 汇总
    all_mids = []
    for p in sorted(page_results.keys()):
        for item in page_results[p]:
            all_mids.append(item.get("upper_mid"))

    print(f"\n阶段一完成！共 {len(page_results)} 页, {len(all_mids)} 个UP主")
    return {"page_results": page_results, "total_pages": total_pages, "all_mids": all_mids}


# ============ 阶段二：抓取UP主详情 ============

async def phase2_fetch_details(context, all_uppers: list):
    """阶段二：逐个访问UP主详情页，拦截API响应
    all_uppers: [(upper_mid, mapping_id), ...] 的列表
    """
    print_banner("阶段二：抓取UP主详情")

    # 加载断点
    ckpt = load_checkpoint(PHASE2_CHECKPOINT)
    detail_results = ckpt.get("detail_data", {})
    completed_mids = set(detail_results.keys())

    uppers_to_fetch = [(mid, mid2) for mid, mid2 in all_uppers if str(mid) not in completed_mids]
    print(f"总计: {len(all_uppers)} 个UP主")
    print(f"已完成: {len(completed_mids)}")
    print(f"本次需抓取: {len(uppers_to_fetch)}")

    if not uppers_to_fetch:
        print("所有UP主详情已抓取完成")
        return detail_results

    print(f"共 {len(uppers_to_fetch)} 个UP主待抓取\n")

    page = await context.new_page()

    # 设置网络拦截，捕获API响应
    captured = {}

    async def handle_response(response):
        url = response.url
        for api_path, key in INTERCEPT_APIS.items():
            if api_path in url:
                try:
                    body = await response.json()
                    if body.get("code") == 0:
                        # 用 key + URL参数 区分不同请求
                        if "trend_type=" in url:
                            trend_type = url.split("trend_type=")[1].split("&")[0]
                            full_key = f"{key}_type{trend_type}"
                        elif "query_type=" in url:
                            query_type = url.split("query_type=")[1].split("&")[0]
                            full_key = f"{key}_query{query_type}"
                        elif "type=" in url and "representative" in url:
                            rep_type = url.split("type=")[1].split("&")[0]
                            full_key = f"{key}_type{rep_type}"
                        else:
                            full_key = key
                        captured[full_key] = body.get("result", {})
                except Exception:
                    pass

    page.on("response", handle_response)

    pbar = tqdm(total=len(uppers_to_fetch), desc="阶段二-详情", unit="人",
                bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")

    for i, (mid, mapping_id) in enumerate(uppers_to_fetch):
        captured.clear()
        detail_url = DETAIL_PAGE_URL.format(mapping_id=mapping_id)

        try:
            # 先跳空白页，强制SPA完整重新加载
            await page.goto("about:blank", wait_until="load", timeout=5000)
            await asyncio.sleep(0.5)
            await page.goto(detail_url, wait_until="networkidle", timeout=30000)
            # 等待页面数据加载
            await asyncio.sleep(2)
        except Exception as e:
            tqdm.write(f"  [mid={mid}] 页面加载异常: {e}")

        # 检查是否触发验证码
        captcha = await page.query_selector(".geetest_panel, .captcha-container, #gc-box")
        if captcha:
            tqdm.write(f"  [mid={mid}] 触发验证码，等待手动验证...")
            save_checkpoint_file(PHASE2_CHECKPOINT, {
                "detail_data": detail_results,
            })
            resolved = await wait_for_captcha_resolve(page, context)
            if not resolved:
                break
            captured.clear()
            try:
                await page.goto("about:blank", wait_until="load", timeout=5000)
                await asyncio.sleep(0.5)
                await page.goto(detail_url, wait_until="networkidle", timeout=30000)
                await asyncio.sleep(2)
            except Exception:
                pass

        # 保存捕获的数据
        if captured:
            detail_results[str(mid)] = dict(captured)
            pbar.set_postfix(成功=len(detail_results), 当前API数=len(captured))
        else:
            tqdm.write(f"  [mid={mid}] 未捕获到数据")

        pbar.update(1)

        # 定期保存断点
        if (i + 1) % SAVE_EVERY == 0:
            save_checkpoint_file(PHASE2_CHECKPOINT, {
                "detail_data": detail_results,
            })


    pbar.close()
    await page.close()

    # 最终保存
    save_checkpoint_file(PHASE2_CHECKPOINT, {
        "detail_data": detail_results,
    })

    print(f"\n阶段二完成！共 {len(detail_results)} 个UP主详情")
    return detail_results


# ============ 数据导出 ============

def extract_price_info(price_infos: list) -> dict:
    result = {"植入视频报价": "", "定制视频报价": "", "直发动态报价": "", "转发动态报价": ""}
    type_map = {1: "植入视频报价", 2: "定制视频报价", 3: "直发动态报价", 4: "转发动态报价"}
    for p in (price_infos or []):
        key = type_map.get(p.get("cooperation_type"))
        if key:
            result[key] = p.get("platform_price", "")
    return result


def extract_distribution(dist_list: list, key_name: str = "section_desc") -> dict:
    """将分布列表转为字典 {描述: 占比}"""
    if not dist_list:
        return {}
    return {item.get(key_name, ""): item.get("count", 0) for item in dist_list}


def build_row(search_item: dict, detail: dict) -> dict:
    """合并列表数据和详情数据，生成一行Excel数据"""
    prices = extract_price_info(search_item.get("price_infos", []))
    draft_data = search_item.get("up_draft_data_view") or {}
    goods_data = search_item.get("goods_up_data_view") or {}

    # 从详情中提取
    portrait = detail.get("portrait", {}) if detail else {}
    # 投稿趋势
    draft_play = detail.get("draft_trend_type3", {}) if detail else {}
    draft_like = detail.get("draft_trend_type4", {}) if detail else {}
    draft_comment = detail.get("draft_trend_type5", {}) if detail else {}
    draft_danmu = detail.get("draft_trend_type6", {}) if detail else {}
    # 粉丝趋势
    fans_total = detail.get("fans_trend_query1", {}) if detail else {}
    fans_inc = detail.get("fans_trend_query2", {}) if detail else {}
    # 代表作品
    personal_works = detail.get("representative_type1", []) if detail else []
    commercial_works = detail.get("representative_type2", []) if detail else []

    row = {
        # ---- 基本信息 (API1) ----
        "UP主昵称": search_item.get("nickname", ""),
        "UP主MID": search_item.get("upper_mid", ""),
        "性别": search_item.get("gender_desc", ""),
        "地区": search_item.get("region_desc", ""),
        "城市": search_item.get("second_region_desc", ""),
        "一级分区": search_item.get("partition_name", ""),
        "二级分区": search_item.get("second_partition_name", ""),
        "MCN公司": search_item.get("mcn_company_name", ""),
        "UP主类型": portrait.get("upper_type_desc", ""),
        "个性签名": portrait.get("signature", ""),

        # ---- 核心指标 (API1) ----
        "粉丝量": search_item.get("fans_num", 0),
        "涨粉量": search_item.get("fans_inc", 0),
        "涨粉率(%)": search_item.get("fans_inc_rate", 0),
        "播放量中位数": search_item.get("average_play_cnt", 0),
        "30天播放量中位数": draft_data.get("play_median_nature_30d", 0),
        "互动率": search_item.get("interactive_rate", 0),
        "热门稿件率": search_item.get("hot_avid_rate", 0),
        "投稿时长": search_item.get("draft_duration", 0),
        "近30天投稿数": search_item.get("avid_cnt_30d", 0),
        "近90天投稿数": search_item.get("avid_cnt_90d", 0),
        "近180天投稿数": search_item.get("avid_cnt_180d", 0),
        "180天热门稿件数": search_item.get("hot_avid_cnt_180d", 0),
        "商业视频数": search_item.get("commercial_avid_cnt", 0),
        "移动端播放占比": search_item.get("app_vv_percent", 0),
        "PC端播放占比": search_item.get("pc_vv_percent", 0),
        "综合评分": search_item.get("synthetical_score", 0),
        "磁力值": search_item.get("magnetic_value", ""),
        "预估播放量": search_item.get("estimated_play", 0),
        "预估花费": search_item.get("estimated_cost", ""),

        # ---- 报价 (API1) ----
        "植入视频报价": prices["植入视频报价"],
        "定制视频报价": prices["定制视频报价"],
        "直发动态报价": prices["直发动态报价"],
        "转发动态报价": prices["转发动态报价"],
        "CPM": search_item.get("cpm", ""),
        "CPC": search_item.get("cpc", ""),

        # ---- 效果数据 (API1) ----
        "自然流量_播放中位数_30天": draft_data.get("play_median_nature_30d", ""),
        "自然流量_CPM_30天": draft_data.get("cpm_nature_30d", ""),
        "自然流量_CPC_30天": draft_data.get("cpc_nature_30d", ""),
        "全部流量_CPM_30天": draft_data.get("cpm_all_30d", ""),
        "全部流量_CPC_30天": draft_data.get("cpc_all_30d", ""),

        # ---- 详情独有-互动均值 (API2) ----
        "平均评论数": portrait.get("average_comment_cnt", ""),
        "平均点赞数": portrait.get("average_like_cnt", ""),
        "平均收藏数": portrait.get("average_collect_cnt", ""),
        "平均弹幕数": portrait.get("average_barrage_cnt", ""),
        "总获赞数": portrait.get("fans_like_num", ""),
        "总视频数": portrait.get("video_num", ""),
        "动态互动中位数": portrait.get("dyn_median_interact_num", ""),
        "动态观看中位数": portrait.get("dyn_median_view_num", ""),

        # ---- 非花火商单 (API2) ----
        "非花火_平均播放量": portrait.get("average_play_cnt_other_h", ""),
        "非花火_平均互动率": portrait.get("average_interactive_rate_other_h", ""),

        # ---- 投稿趋势汇总 (API3) ----
        "播放量_中位数_趋势": draft_play.get("median", ""),
        "播放量_最高": draft_play.get("max_cnt", ""),
        "播放量_最低": draft_play.get("min_cnt", ""),
        "点赞量_中位数_趋势": draft_like.get("median", ""),
        "评论量_中位数_趋势": draft_comment.get("median", ""),
        "弹幕量_中位数_趋势": draft_danmu.get("median", ""),

        # ---- 粉丝趋势 (API4) ----
        "7天涨粉量": fans_total.get("fans_inc7", ""),
        "30天涨粉量": fans_total.get("fans_inc30", ""),
        "90天涨粉量": fans_total.get("fans_inc90", ""),
        "180天涨粉量": fans_total.get("fans_inc180", ""),
        "365天涨粉量": fans_total.get("fans_inc365", ""),
        "7天涨粉率(%)": fans_total.get("fans_inc_rate7", ""),
        "30天涨粉率(%)": fans_total.get("fans_inc_rate30", ""),
        "90天涨粉率(%)": fans_total.get("fans_inc_rate90", ""),
        "180天涨粉率(%)": fans_total.get("fans_inc_rate180", ""),
        "365天涨粉率(%)": fans_total.get("fans_inc_rate365", ""),

        # ---- 代表作品 (API5) ----
        "商业作品数": len(commercial_works) if isinstance(commercial_works, list) else 0,
        "个人作品数": len(personal_works) if isinstance(personal_works, list) else 0,
    }

    # ---- 关注用户画像 (API2 完整分布) ----
    sax = extract_distribution(portrait.get("sax_distributions", []))
    for desc, val in sax.items():
        row[f"关注用户_{desc}占比(%)"] = val

    age = extract_distribution(portrait.get("age_distributions", []))
    for desc, val in age.items():
        row[f"关注用户_{desc}占比(%)"] = val

    # ---- 观看用户画像 (API2 完整分布) ----
    sax_a = extract_distribution(portrait.get("sax_distributions_audience", []))
    for desc, val in sax_a.items():
        row[f"观看用户_{desc}占比(%)"] = val

    age_a = extract_distribution(portrait.get("age_distributions_audience", []))
    for desc, val in age_a.items():
        row[f"观看用户_{desc}占比(%)"] = val

    fans_vv = extract_distribution(portrait.get("fans_vv_distributions_audience", []))
    for desc, val in fans_vv.items():
        row[f"观看用户_{desc}占比(%)"] = val

    # ---- 城市等级分布 (API2) ----
    city = extract_distribution(portrait.get("city_distributions", []))
    for desc, val in city.items():
        row[f"关注用户_城市_{desc}(%)"] = val

    # ---- 设备分布 Top3 (API2) ----
    device = extract_distribution(portrait.get("device_distributions", []))
    for j, (desc, val) in enumerate(sorted(device.items(), key=lambda x: -x[1])[:3]):
        row[f"关注用户_设备Top{j+1}"] = f"{desc}({val}%)"

    # ---- 地区Top5 (API2) ----
    region = extract_distribution(portrait.get("top_region_distributions", []))
    for j, (desc, val) in enumerate(sorted(region.items(), key=lambda x: -x[1])[:5]):
        row[f"关注用户_地区Top{j+1}"] = f"{desc}({val}%)"

    # ---- 品类偏好 (API2) ----
    row["关注用户_品类偏好"] = ", ".join(portrait.get("first_categories_profile", []) or [])
    row["观看用户_品类偏好"] = ", ".join(portrait.get("first_categories_profile_audience", []) or [])
    row["适合投放品类"] = ", ".join(portrait.get("category_names", []) or [])

    # ---- 带货 (API1) ----
    row["带货等级"] = goods_data.get("goods_up_level", "")
    row["带货品类"] = ", ".join(goods_data.get("goods_category", []) or [])
    row["带货GPM"] = goods_data.get("goods_gpm", "")

    # ---- 链接 ----
    row["头像URL"] = search_item.get("head_img", "")
    row["B站主页"] = f"https://space.bilibili.com/{search_item.get('upper_mid', '')}"
    row["火花主页"] = DETAIL_PAGE_URL.format(mapping_id=search_item.get("mapping_id", ""))

    return row


def save_to_excel(rows: list[dict], filename: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UP主数据"

    if not rows:
        print("没有数据可保存")
        return

    headers = list(rows[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    for row_idx, row_data in enumerate(rows, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col, value=row_data.get(header, ""))

    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(len(str(header)) * 2 + 4, 12)

    wb.save(filename)
    print(f"数据已保存到: {filename}")


# ============ 主流程 ============

def check_debug_port() -> bool:
    """检查Chrome调试端口是否可用"""
    import urllib.request
    try:
        urllib.request.urlopen(CDP_URL + "/json", timeout=2)
        return True
    except Exception:
        return False


def wait_for_chrome():
    """等待用户手动启动Chrome调试模式"""
    if check_debug_port():
        print("[状态] Chrome调试端口已连接")
        return True

    print("\n[状态] 未检测到Chrome调试端口")
    print()
    print("请选择启动方式:")
    print("  a. 自动启动（脚本帮你关闭Chrome并重启）")
    print("  b. 手动启动（你自己在终端中运行命令）")
    print()
    mode = input("请输入 a 或 b: ").strip().lower()

    if mode == "a":
        import subprocess as sp
        print("\n正在关闭所有Chrome进程...")
        sp.run(["taskkill", "/F", "/IM", "chrome.exe"],
               stdout=sp.DEVNULL, stderr=sp.DEVNULL)
        time.sleep(3)
        # 用bat文件启动（避免路径转义问题）
        bat_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "start_chrome.bat")
        if not os.path.exists(bat_path):
            with open(bat_path, "w") as f:
                f.write('@echo off\n')
                f.write('taskkill /F /IM chrome.exe >nul 2>&1\n')
                f.write('ping -n 4 127.0.0.1 >nul\n')
                f.write('start "" "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"'
                        ' --remote-debugging-port=9222'
                        ' --user-data-dir="C:\\Users\\pc\\AppData\\Local\\Google\\ChromeDebug"\n')
        os.system(f'"{bat_path}"')
        print("Chrome已启动，等待就绪...")
        time.sleep(5)
    else:
        print()
        print("请在终端中运行（注意必须在同一行）:")
        print()
        print('  "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"'
              ' --remote-debugging-port=9222'
              ' --user-data-dir="C:\\Users\\pc\\AppData\\Local\\Google\\ChromeDebug"')
        print()
        print("  或者直接双击项目目录下的 start_chrome.bat")
        print()
        print("  注意: 必须使用独立目录(ChromeDebug)，不能用默认目录")
        print("  首次使用需要在Chrome中登录 https://huahuo.bilibili.com")
        print("  之后会自动保留登录态")
        print()
        input("启动完成后按回车继续...")
    print()
    input("完成后按回车继续...")

    # 等待端口
    print("检测Chrome调试端口", end="", flush=True)
    for i in range(30):
        if check_debug_port():
            print(" 已连接！")
            return True
        time.sleep(1)
        print(".", end="", flush=True)

    print("\n\n[错误] 仍无法连接调试端口")
    print("请确认:")
    print("  - 启动前已关闭所有Chrome（包括后台进程）")
    print("  - 命令中包含 --remote-debugging-port=9222")
    print("  - 访问 http://localhost:9222 有响应")
    return False


async def verify_login(context) -> bool:
    """验证火花平台登录态是否有效"""
    print("[状态] 验证登录态...", end=" ", flush=True)
    page = await context.new_page()
    try:
        url = f"{SEARCH_URL}?ct_id={SEARCH_CT_ID}&page=1"
        await page.goto(url, wait_until="networkidle", timeout=15000)
        text = await page.evaluate("() => document.body.innerText")
        body = json.loads(text)

        if body.get("code") == 0 and isinstance(body.get("result"), dict):
            count = body["result"].get("total_count", 0)
            print(f"成功（共{count}条数据）")
            return True
        elif body.get("code") == 1031:
            print("触发频率限制，请稍等几分钟后重试")
            return False
        else:
            print(f"失败（code={body.get('code')}, msg={body.get('msg', '')}）")
            print("  请在Chrome中打开火花平台确认登录状态")
            return False
    except Exception as e:
        print(f"失败（{e}）")
        return False
    finally:
        await page.close()


async def main():
    print_banner("B站火花平台UP主数据抓取（浏览器版）")

    ensure_dir()

    # 步骤1：连接Chrome
    if not wait_for_chrome():
        return

    async with async_playwright() as p:
        try:
            browser = await p.chromium.connect_over_cdp(CDP_URL)
        except Exception as e:
            print(f"\n[错误] 连接Chrome失败: {e}")
            return

        context = browser.contexts[0]
        print("[状态] 已连接到Chrome浏览器")

        # 步骤2：验证登录态
        if not await verify_login(context):
            print("\n请在Chrome中登录火花平台后重新运行脚本")
            return

        print()  # 空行后进入菜单

        # 主菜单
        while True:
            print("-" * 40)
            print("  请选择操作:")
            print("  1. 抓取UP主列表（500页总览数据）")
            print("  2. 抓取UP主详情（需先完成步骤1）")
            print("  3. 导出Excel（用已有数据导出）")
            print("  4. 查看当前进度")
            print("  0. 退出")
            print("-" * 40)

            choice = input("请输入选项: ").strip()

            if choice == "1":
                await phase1_fetch_list(context)

            elif choice == "2":
                # 检查阶段一是否有数据
                ckpt1 = load_checkpoint(PHASE1_CHECKPOINT)
                if not ckpt1 or not ckpt1.get("page_data"):
                    print("\n错误: 请先执行步骤1抓取UP主列表\n")
                    continue

                # 汇总所有 (upper_mid, mapping_id) 配对
                all_uppers = []
                for pg in sorted(ckpt1["page_data"].keys(), key=int):
                    for item in ckpt1["page_data"][pg]:
                        all_uppers.append((item.get("upper_mid"), item.get("mapping_id")))

                ckpt2 = load_checkpoint(PHASE2_CHECKPOINT)
                done_count = len(ckpt2.get("detail_data", {}))

                print(f"\n共 {len(all_uppers)} 个UP主，已完成详情: {done_count}")
                print("  输入 y 或回车 — 抓取全部")
                print("  输入 数字N — 只抓取前N个")
                print("  输入 n — 取消")
                sub = input("请选择: ").strip().lower()

                if sub == "n":
                    continue
                elif sub.isdigit():
                    n = int(sub)
                    await phase2_fetch_details(context, all_uppers[:n])
                else:
                    await phase2_fetch_details(context, all_uppers)

            elif choice == "3":
                # 导出Excel
                ckpt1 = load_checkpoint(PHASE1_CHECKPOINT)
                if not ckpt1 or not ckpt1.get("page_data"):
                    print("\n错误: 没有列表数据，请先执行步骤1\n")
                    continue

                ckpt2 = load_checkpoint(PHASE2_CHECKPOINT)
                detail_results = ckpt2.get("detail_data", {})

                print_banner("导出数据")
                all_rows = []
                for pg in sorted(ckpt1["page_data"].keys(), key=int):
                    for item in ckpt1["page_data"][pg]:
                        mid = str(item.get("upper_mid", ""))
                        detail = detail_results.get(mid, {})
                        all_rows.append(build_row(item, detail))

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"huohua_full_{timestamp}.xlsx"
                save_to_excel(all_rows, filename)

                print(f"总计: {len(all_rows)} 条数据")
                print(f"其中有详情数据: {len(detail_results)} 条")
                print(f"文件: {filename}\n")

            elif choice == "4":
                ckpt1 = load_checkpoint(PHASE1_CHECKPOINT)
                ckpt2 = load_checkpoint(PHASE2_CHECKPOINT)
                p1_pages = len(ckpt1.get("page_data", {}))
                p1_total = ckpt1.get("total_pages", 0)
                p1_mids = sum(len(v) for v in ckpt1.get("page_data", {}).values())
                p2_done = len(ckpt2.get("detail_data", {}))

                print(f"\n  列表进度: {p1_pages}/{p1_total} 页, {p1_mids} 个UP主")
                print(f"  详情进度: {p2_done}/{p1_mids} 个UP主")
                if ckpt1.get("page_data"):
                    print(f"  列表数据: 有 (可导出)")
                if p2_done > 0:
                    print(f"  详情数据: 有 ({p2_done}条)")
                print()

            elif choice == "0":
                print("退出")
                break
            else:
                print("无效选项，请重新输入\n")


if __name__ == "__main__":
    asyncio.run(main())
